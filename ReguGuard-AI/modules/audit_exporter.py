import io
import re
from datetime import date

from PyPDF2 import PdfReader


def _normalize_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def extract_noncompliant_line(full_text: str, finding: dict) -> str:
    """
    Best-effort: return ONE specific line/sentence from the uploaded document that
    most likely corresponds to the non-compliance finding.

    Since full-document analysis doesn't always return an exact pinpoint, we select
    the best matching line using the finding's area + issue keywords.
    """
    if not full_text:
        return ""

    area = str(finding.get("area", "") or "")
    issue = str(finding.get("issue", "") or "")

    area_words = [w for w in re.split(r"[/\s]+", _normalize_ws(area).lower()) if len(w) >= 3]
    issue_words = [w for w in re.split(r"\W+", _normalize_ws(issue).lower()) if len(w) >= 4]
    issue_words = issue_words[:12]

    # IMPORTANT: we must return an EXACT line from the SOP (verbatim),
    # so we ONLY consider existing line breaks from the extracted text.
    raw_lines = [ln.rstrip("\r") for ln in full_text.splitlines() if ln.strip()]

    # Score each candidate line
    best_line = ""
    best_score = -10**9
    for ln in raw_lines[:2000]:
        ln_norm = _normalize_ws(ln).lower()
        if len(ln_norm) < 12:
            continue
        score = 0
        score += sum(3 for w in area_words if w in ln_norm)
        score += sum(2 for w in issue_words if w in ln_norm)
        # Prefer "weak policy" language that often indicates non-compliance
        if any(x in ln_norm for x in ["should", "may", "when possible", "if feasible", "recommended"]):
            score += 2
        # Slight penalty for very long lines
        score -= max(0, (len(ln_norm) - 260) // 40)
        if score > best_score:
            best_score = score
            best_line = ln
        # Early stop if we have a very strong match
        if best_score >= 12:
            break

    # Return verbatim line (no rewriting, no trimming beyond removing trailing CR)
    if not best_line:
        return ""
    return best_line


def _classify_issue_from_line(line_text: str) -> tuple[str, str]:
    """
    Strict, text-only classification (no inference).
    Returns (issue_type, severity).
    Severity must be one of:
      - Critical GMP Violation
      - Major Gap
      - Documentation Weakness
      - Observation
    """
    ln = (line_text or "").lower()
    weak = ["should", "may", "when possible", "if feasible", "recommended", "as appropriate", "where practicable"]
    outdated = ["superseded", "replaced by", "no longer", "deprecated", "legacy", "previous version", "old standard", "former"]
    explicit_bad = ["not required", "ignore", "do not", "without approval", "no need to", "skip"]

    if any(x in ln for x in explicit_bad):
        return "Explicit policy conflict", "Critical GMP Violation"
    if any(x in ln for x in outdated):
        return "Outdated reference", "Major Gap"
    if any(x in ln for x in weak):
        return "Documentation Clarity Observation", "Observation"
    # If we can't point to an explicit problem on the line itself, don't claim a violation
    return "", ""


def find_pdf_page_number(pdf_bytes: bytes, excerpt: str) -> int | None:
    """
    Best-effort: return 1-based page number where excerpt appears in the PDF.
    Returns None if not found or not applicable.
    """
    if not pdf_bytes or not excerpt:
        return None

    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
    except Exception:
        return None

    needle = _normalize_ws(excerpt).lower()
    if len(needle) > 120:
        needle = needle[:120]

    for i, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text() or ""
        except Exception:
            continue
        hay = _normalize_ws(page_text).lower()
        if not hay:
            continue
        if needle and needle in hay:
            return i + 1

    # fallback: token-based
    tokens = [t for t in re.split(r"\W+", needle) if len(t) >= 5][:6]
    if not tokens:
        return None
    for i, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text() or ""
        except Exception:
            continue
        hay = _normalize_ws(page_text).lower()
        if sum(1 for t in tokens if t in hay) >= max(2, len(tokens) // 2):
            return i + 1
    return None


def pick_source_url_for_finding(finding: dict, citations: list[dict]) -> str:
    """
    Pick one specific source link (prefer ones matching applicable_regulations;
    otherwise highest relevance_score).
    """
    if not citations:
        return ""

    applicable = [str(x).lower() for x in (finding.get("applicable_regulations") or [])]
    if applicable:
        for cit in citations:
            title = str(cit.get("title", "")).lower()
            if any(a and a in title for a in applicable):
                return str(cit.get("url", "") or "")

    best = max(citations, key=lambda c: float(c.get("relevance_score", 0) or 0))
    return str(best.get("url", "") or "")


def _find_page_in_page_texts(page_texts: list[str], excerpt: str) -> int | None:
    needle = _normalize_ws(excerpt).lower()
    if len(needle) > 120:
        needle = needle[:120]
    for i, hay in enumerate(page_texts):
        if needle and needle in hay:
            return i + 1
    tokens = [t for t in re.split(r"\W+", needle) if len(t) >= 5][:6]
    if not tokens:
        return None
    for i, hay in enumerate(page_texts):
        if sum(1 for t in tokens if t in hay) >= max(2, len(tokens) // 2):
            return i + 1
    return None


def build_audit_excel(
    *,
    findings: list[dict],
    full_text: str,
    uploaded_filename: str | None,
    uploaded_file_bytes: bytes | None,
    citations: list[dict],
    approved_changes: dict,
) -> bytes:
    """
    Build an Excel file with:
    A: Non-compliant line/excerpt
    B: Page number (best-effort; only reliable for PDF)
    C: Source link
    D: Action (Approved / Not approved)
    E: Date
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise RuntimeError("Excel export requires 'openpyxl'. Install dependencies and restart.") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Audit Log"

    headers = [
        "Non-compliant line",
        "Page number",
        "Source link",
        "Action",
        "Date",
        "Severity",
        "Suggestion",
    ]
    ws.append(headers)

    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="7C3AED")  # violet
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    today = date.today().isoformat()
    is_pdf = bool(uploaded_filename and uploaded_filename.lower().endswith(".pdf"))
    page_texts = None
    if is_pdf and uploaded_file_bytes:
        try:
            reader = PdfReader(io.BytesIO(uploaded_file_bytes))
            page_texts = []
            for p in reader.pages:
                try:
                    page_texts.append(_normalize_ws((p.extract_text() or "")).lower())
                except Exception:
                    page_texts.append("")
        except Exception:
            page_texts = None

    def _severity_bucket(finding: dict) -> str:
        # Requested scale: low / medium / high / extreme
        rl = str(finding.get("risk_level", "")).strip().lower()
        st = str(finding.get("status", "")).strip().lower()
        if rl == "high" and st in ("non_compliant", "outdated"):
            return "Extreme"
        if rl == "high":
            return "High"
        if rl == "medium":
            return "Medium"
        if rl == "low":
            return "Low"
        return "Medium"

    severity_rank = {"Extreme": 4, "High": 3, "Medium": 2, "Low": 1}

    rows = []
    for idx, f in enumerate(findings or []):
        status = str(f.get("status", "")).strip().lower()
        if status == "compliant":
            continue

        line_text = extract_noncompliant_line(full_text, f)
        if not line_text:
            # Cannot quote an exact line → do not flag
            continue
        page_no = _find_page_in_page_texts(page_texts, line_text) if (is_pdf and page_texts) else None
        source_url = pick_source_url_for_finding(f, citations)

        area = str(f.get("area", "Section"))
        approved_key = f"{area}_{idx}"
        action = "Approved" if approved_key in (approved_changes or {}) else "Not approved"

        severity = _severity_bucket(f)
        suggestion = str(f.get("suggestion", "") or "").strip()
        if source_url and suggestion:
            suggestion = f"{suggestion}\nSource: {source_url}"
        elif source_url and not suggestion:
            suggestion = f"Source: {source_url}"

        rows.append((severity_rank.get(severity, 2), [line_text, page_no or "N/A", source_url, action, today, severity, suggestion]))

    # Sort by decreasing severity
    rows.sort(key=lambda x: x[0], reverse=True)
    for _, r in rows:
        ws.append(r)

    # If nothing explicit could be quoted, add a single row per requirement
    if ws.max_row == 1:
        ws.append(["No explicit compliance violation detected.", "N/A", "", "Not approved", today, "Low", ""])

    # Column widths + wrapping
    col_widths = [60, 14, 44, 14, 12, 10, 60]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_strict_audit_csv_from_analysis(
    *,
    findings: list[dict],
    full_text: str,
    citations: list[dict],
    approved_changes: dict,
) -> bytes:
    """
    Build STRICT CSV (for Excel) with four columns:
      1) Exact full line from SOP (verbatim; must exist as a line in extracted text)
      2) Issue type (includes Approved/Not approved)
      3) Severity (Critical GMP Violation | Major Gap | Documentation Weakness | Observation)
      4) Suggestion (dynamic; include one source URL for traceability)

    Rules:
    - Only flag if we can quote an exact line as-is (verbatim).
    - No inference: issue type / severity derived from the line text itself.
    - If no explicit non-compliance is found, output a single row:
      "No explicit compliance violation detected."
    """
    import csv

    lines = [ln.rstrip("\r") for ln in (full_text or "").splitlines()]
    line_set = set(lines)

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["SOP Line (verbatim)", "Issue Type", "Severity", "Suggestion"])

    wrote_any = False
    for idx, f in enumerate(findings or []):
        status = str(f.get("status", "")).strip().lower()
        if status == "compliant":
            continue

        line_text = extract_noncompliant_line(full_text, f)
        if not line_text:
            continue
        # Enforce exact line match from extracted SOP text
        if line_text not in line_set:
            continue

        issue_type, severity = _classify_issue_from_line(line_text)
        if not issue_type or not severity:
            # If we can't explicitly classify from the line, do not flag it
            continue

        area = str(f.get("area", "Section"))
        approved_key = f"{area}_{idx}"
        action = "Approved" if approved_key in (approved_changes or {}) else "Not approved"
        issue_type = f"{issue_type} — {action}"

        suggestion = str(f.get("suggestion", "") or "").strip()
        source_url = pick_source_url_for_finding(f, citations)
        if source_url:
            if suggestion:
                suggestion = f"{suggestion}\nSource: {source_url}"
            else:
                suggestion = f"Source: {source_url}"

        writer.writerow([line_text, issue_type, severity, suggestion])
        wrote_any = True

    if not wrote_any:
        writer.writerow(["No explicit compliance violation detected.", "Observation", "Observation", ""])

    return out.getvalue().encode("utf-8")

