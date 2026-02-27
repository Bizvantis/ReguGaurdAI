import streamlit as st
import os
import json
import time
import re
from datetime import datetime
import pandas as pd
from io import BytesIO

from modules.sop_parser import (
    extract_text_from_file,
    extract_text_and_pages_from_file,
    extract_text_from_bytes,
)
from modules.regulation_scraper import scrape_regulations, get_regulation_sources
from modules.rag_engine import RAGEngine
from modules.compliance_analyzer import (
    get_grok_client,
    is_document_sop,
    analyze_full_document_compliance,
)
from modules.document_editor import build_updated_document, build_updated_text_with_additions
from modules.domain_classifier import detect_document_domain
from modules.history_manager import save_analysis_history, load_analysis_history, get_analysis_by_id, delete_analysis

st.set_page_config(
    page_title="ReguGuard AI — Compliance Manager",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Modern, colourful & aesthetic UI — vibrant yet professional
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap');
    :root {
        --rg-teal: #0d9488;
        --rg-violet: #7c3aed;
        --rg-amber: #f59e0b;
        --rg-rose: #f43f5e;
        --rg-emerald: #10b981;
        --rg-sky: #0ea5e9;
        --rg-indigo: #6366f1;
        --rg-mint: #99f6e4;
        --rg-cream: #fffbeb;
        --rg-slate: #1e293b;
        --rg-light: #f8fafc;
    }
    .stApp {
        background: linear-gradient(165deg, #f0f9ff 0%, #faf5ff 35%, #fefce8 70%, #f0fdf4 100%);
        background-attachment: fixed;
    }
    .main .block-container { padding-top: 1.5rem; padding-bottom: 3rem; max-width: 1400px; }
    h1, h2, h3 { font-family: 'Plus Jakarta Sans', system-ui, sans-serif; font-weight: 700; color: var(--rg-slate); }
    .rg-header {
        background: linear-gradient(135deg, var(--rg-teal) 0%, var(--rg-violet) 50%, var(--rg-indigo) 100%);
        padding: 1.5rem 1.75rem; border-radius: 16px; margin-bottom: 1.75rem;
        box-shadow: 0 10px 40px -10px rgba(13, 148, 136, 0.35), 0 4px 20px -5px rgba(124, 58, 237, 0.25);
    }
    .rg-header h1 { font-size: 1.65rem; font-weight: 700; letter-spacing: -0.02em; color: #fff; margin: 0; text-shadow: 0 1px 2px rgba(0,0,0,0.1); }
    .rg-header .tagline { font-size: 0.875rem; color: rgba(255,255,255,0.9); margin-top: 0.35rem; font-weight: 500; }
    .rg-header .shield { font-size: 1.5rem; margin-right: 0.5rem; vertical-align: middle; }
    .rg-stepper {
        display: flex; align-items: center; gap: 0; margin-bottom: 1.75rem; flex-wrap: wrap;
        background: rgba(255,255,255,0.85); backdrop-filter: blur(12px); padding: 0.9rem 1.25rem;
        border-radius: 12px; border: 1px solid rgba(255,255,255,0.8); box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    }
    .rg-step {
        display: flex; align-items: center; font-size: 0.8125rem; color: #94a3b8; font-weight: 500;
        font-family: 'Plus Jakarta Sans', system-ui, sans-serif;
    }
    .rg-step.active { color: var(--rg-violet); font-weight: 700; }
    .rg-step.done { color: var(--rg-emerald); font-weight: 600; }
    .rg-step-dot {
        width: 10px; height: 10px; border-radius: 50%; background: #cbd5e1; margin-right: 0.5rem;
        transition: all 0.25s ease;
    }
    .rg-step.active .rg-step-dot {
        background: linear-gradient(135deg, var(--rg-violet), var(--rg-indigo));
        box-shadow: 0 0 0 3px rgba(124, 58, 237, 0.25), 0 2px 8px rgba(124, 58, 237, 0.3);
    }
    .rg-step.done .rg-step-dot {
        background: linear-gradient(135deg, var(--rg-emerald), #34d399);
        box-shadow: 0 2px 6px rgba(16, 185, 129, 0.35);
    }
    .rg-step-sep {
        width: 28px; height: 2px; margin: 0 6px;
        background: linear-gradient(90deg, #e2e8f0, #cbd5e1); border-radius: 1px;
    }
    .rg-card {
        background: rgba(255,255,255,0.9); backdrop-filter: blur(10px);
        border: 1px solid rgba(255,255,255,0.9); border-radius: 12px;
        padding: 1.25rem; margin-bottom: 1rem;
        box-shadow: 0 4px 24px -4px rgba(0,0,0,0.08), 0 2px 8px -2px rgba(0,0,0,0.04);
    }
    .rg-metric {
        background: rgba(255,255,255,0.95); backdrop-filter: blur(10px);
        border: 1px solid rgba(255,255,255,0.9); border-radius: 14px;
        padding: 1.15rem 1.25rem; text-align: center;
        box-shadow: 0 4px 20px -4px rgba(0,0,0,0.08), 0 2px 10px -2px rgba(0,0,0,0.04);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .rg-metric:hover { transform: translateY(-2px); box-shadow: 0 8px 30px -6px rgba(0,0,0,0.12); }
    .rg-metric .value { font-size: 1.6rem; font-weight: 700; color: var(--rg-slate); line-height: 1.2; font-family: 'Plus Jakarta Sans', system-ui, sans-serif; }
    .rg-metric .label { font-size: 0.7rem; color: #64748b; text-transform: uppercase; letter-spacing: 0.06em; margin-top: 0.3rem; font-weight: 600; }
    .rg-metric.score-good .value { color: var(--rg-emerald); }
    .rg-metric.score-mid .value { color: var(--rg-amber); }
    .rg-metric.score-low .value { color: var(--rg-rose); }
    .rg-metric.accent-teal { border-top: 3px solid var(--rg-teal); }
    .rg-metric.accent-violet { border-top: 3px solid var(--rg-violet); }
    .rg-metric.accent-amber { border-top: 3px solid var(--rg-amber); }
    .rg-badge {
        font-size: 0.6875rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em;
        padding: 0.35rem 0.65rem; border-radius: 8px; display: inline-block;
        box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    }
    .rg-badge.compliant { background: linear-gradient(135deg, #d1fae5, #a7f3d0); color: #065f46; }
    .rg-badge.partial { background: linear-gradient(135deg, #fef3c7, #fde68a); color: #92400e; }
    .rg-badge.non-compliant { background: linear-gradient(135deg, #fecaca, #fca5a5); color: #991b1b; }
    .rg-badge.outdated { background: linear-gradient(135deg, #e2e8f0, #cbd5e1); color: #475569; }
    .rg-risk-high { color: var(--rg-rose); font-weight: 700; font-size: 0.75rem; }
    .rg-risk-medium { color: var(--rg-amber); font-weight: 700; font-size: 0.75rem; }
    .rg-risk-low { color: var(--rg-emerald); font-weight: 700; font-size: 0.75rem; }
    .rg-citation {
        background: linear-gradient(135deg, rgba(14, 165, 233, 0.08), rgba(99, 102, 241, 0.08));
        border: 1px solid rgba(14, 165, 233, 0.2); border-left: 4px solid var(--rg-sky);
        border-radius: 10px; padding: 0.85rem 1rem; margin: 0.5rem 0; font-size: 0.8125rem;
        box-shadow: 0 2px 8px rgba(14, 165, 233, 0.06);
    }
    .rg-citation a { color: var(--rg-violet); text-decoration: none; font-weight: 600; }
    .rg-citation a:hover { text-decoration: underline; }
    .rg-suggestion {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.12), rgba(52, 211, 153, 0.08));
        border: 1px solid rgba(16, 185, 129, 0.35); border-left: 4px solid var(--rg-emerald);
        border-radius: 10px; padding: 1rem 1.1rem; margin: 0.5rem 0; font-size: 0.875rem; color: #065f46;
        font-weight: 500; box-shadow: 0 2px 10px rgba(16, 185, 129, 0.08);
    }
    .rg-divider {
        height: 2px; margin: 1.25rem 0;
        background: linear-gradient(90deg, transparent, rgba(148, 163, 184, 0.4), transparent);
        border: none;
    }
    .rg-upload-zone {
        border: 2px dashed var(--rg-teal); border-radius: 14px; padding: 2rem; text-align: center;
        background: linear-gradient(135deg, rgba(13, 148, 136, 0.06), rgba(124, 58, 237, 0.04));
        transition: all 0.3s ease;
    }
    .rg-upload-zone:hover { border-color: var(--rg-violet); background: linear-gradient(135deg, rgba(124, 58, 237, 0.08), rgba(99, 102, 241, 0.06)); box-shadow: 0 8px 24px rgba(124, 58, 237, 0.12); }
    .rg-section-title {
        font-family: 'Plus Jakarta Sans', system-ui, sans-serif; font-weight: 700; color: var(--rg-slate);
        font-size: 1.1rem; margin-bottom: 0.25rem;
    }
    .rg-section-caption { color: #64748b; font-size: 0.875rem; margin-bottom: 1rem; }
    .rg-hero-metric {
        background: linear-gradient(135deg, rgba(255,255,255,0.95), rgba(248,250,252,0.9));
        border-radius: 16px; padding: 1.25rem; text-align: center;
        border: 1px solid rgba(255,255,255,0.9);
        box-shadow: 0 8px 32px -8px rgba(13, 148, 136, 0.2), 0 4px 16px -4px rgba(0,0,0,0.06);
    }
    div[data-testid="stSidebar"] {
        background: linear-gradient(180deg, rgba(248,250,252,0.98) 0%, rgba(241,245,249,0.98) 100%);
        border-right: 1px solid rgba(148, 163, 184, 0.2);
    }
    div[data-testid="stSidebar"] .stMarkdown { font-family: 'Plus Jakarta Sans', system-ui, sans-serif; }
    .rg-sidebar-section { font-size: 0.8125rem; color: #64748b; margin-top: 0.5rem; }
    .rg-sidebar-section h4 { font-size: 0.6875rem; text-transform: uppercase; letter-spacing: 0.06em; color: var(--rg-violet); font-weight: 700; margin-bottom: 0.5rem; }
    div[data-testid="stFileUploader"] {
        background: linear-gradient(135deg, rgba(13, 148, 136, 0.06), rgba(124, 58, 237, 0.04));
        border: 2px dashed rgba(13, 148, 136, 0.4);
        border-radius: 14px; padding: 1.5rem;
    }
    div[data-testid="stFileUploader"]:hover { border-color: var(--rg-teal); background: rgba(13, 148, 136, 0.08); }
    .stProgress > div > div { background: linear-gradient(90deg, var(--rg-teal), var(--rg-violet)); }

    /* Base button styling — soft, rounded, human */
    div.stButton > button {
        border-radius: 999px;
        border: 1px solid #c4b5fd;        /* soft violet border */
        background: #e0e7ff;              /* light indigo/violet */
        color: #111827;
        font-weight: 500;
        padding: 0.45rem 1.3rem;
        box-shadow: 0 1px 3px rgba(15,23,42,0.08);
        transition: background 0.15s ease, box-shadow 0.15s ease, transform 0.1s ease;
    }
    div.stButton > button:hover {
        background: #e5e7eb;              /* soft neutral on hover */
        box-shadow: 0 2px 6px rgba(15,23,42,0.18);
        transform: translateY(-1px);
    }

    /* Domain-specific approve/reject styling (using markers before the button) */
    .rg-approve-marker + div.stButton > button {
        background: #22c55e; /* green */
        border-color: #16a34a;
        color: #f0fdf4;
        box-shadow: 0 1px 4px rgba(22,163,74,0.25);
    }
    .rg-approve-marker + div.stButton > button:hover {
        background: #ffffff; /* turn white on hover */
        border-color: #16a34a;
        color: #166534;
        box-shadow: 0 3px 8px rgba(22,163,74,0.35);
    }
    .rg-reject-marker + div.stButton > button {
        background: #ef4444; /* red */
        border-color: #b91c1c;
        color: #fef2f2;
        box-shadow: 0 1px 4px rgba(185,28,28,0.25);
    }
    .rg-reject-marker + div.stButton > button:hover {
        background: #ffffff; /* turn white on hover */
        border-color: #b91c1c;
        color: #b91c1c;
        box-shadow: 0 3px 8px rgba(185,28,28,0.35);
    }
    
    /* History sidebar styling */
    div[data-testid="stSidebar"] button[key^="history_"] {
        text-align: left;
        padding: 0.6rem 0.9rem;
        margin-bottom: 0.4rem;
        border-radius: 8px;
        font-size: 0.85rem;
        background: rgba(255,255,255,0.7);
        border: 1px solid rgba(148,163,184,0.2);
        transition: all 0.2s ease;
    }
    div[data-testid="stSidebar"] button[key^="history_"]:hover {
        background: rgba(241,245,249,0.9);
        border-color: rgba(124,58,237,0.3);
        transform: translateX(2px);
    }
    
    /* Navigation button styling */
    div[data-testid="stSidebar"] [key*="btn_"] button {
        border-radius: 10px !important;
        border: 1.5px solid rgba(124,58,237,0.2) !important;
        background: rgba(255,255,255,0.6) !important;
        color: #64748b !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
        padding: 0.65rem 0.75rem !important;
        transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }
    
    div[data-testid="stSidebar"] [key*="btn_"] button:hover {
        background: rgba(124,58,237,0.08) !important;
        border-color: rgba(124,58,237,0.4) !important;
        color: #7c3aed !important;
        box-shadow: 0 4px 12px rgba(124,58,237,0.15) !important;
        transform: translateY(-2px) !important;
    }
    
    div[data-testid="stSidebar"] [key*="btn_"] button:active {
        transform: translateY(0) !important;
    }

</style>
""", unsafe_allow_html=True)


def init_session_state():
    defaults = {
        "sop_text": None,
        "sop_pages": None,
        "uploaded_file_bytes": None,
        "uploaded_file_name": None,
        "last_updated_sop_bytes": None,
        "last_updated_sop_name": None,
        "last_updated_sop_mime": None,
        "last_workflow_step": "upload",
        "nav_page": "Workflow",
        "is_sop_result": None,
        "regulations": [],
        "rag_engine": None,
        "full_document_analysis": None,
        "approved_changes": {},
        "rejected_changes": {},
        "analysis_complete": False,
        "scraping_complete": False,
        "current_step": "upload",
        "document_domain": None,
        "document_domain_info": None,
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


init_session_state()


def render_header():
    st.markdown("""
    <div class="rg-header">
        <h1><span class="shield">🛡️</span> ReguGuard AI</h1>
        <p class="tagline">Autonomous compliance manager for SMEs · Regulatory synchronization</p>
    </div>
    """, unsafe_allow_html=True)


def render_stepper(current_step):
    steps = ["upload", "check_document", "scrape", "analyze", "results"]
    labels = ["Upload", "Check", "Regulations", "Analysis", "Results"]
    idx = steps.index(current_step) if current_step in steps else 0
    parts = []
    for i, (s, label) in enumerate(zip(steps, labels)):
        cls = "rg-step"
        if i == idx:
            cls += " active"
        elif i < idx:
            cls += " done"
        parts.append(f'<span class="{cls}"><span class="rg-step-dot"></span>{label}</span>')
        if i < len(steps) - 1:
            parts.append('<span class="rg-step-sep"></span>')
    st.markdown(f'<div class="rg-stepper">{"".join(parts)}</div>', unsafe_allow_html=True)


def render_sidebar():
    with st.sidebar:
        # Navigation heading at the top
        st.markdown("""
        <div style="
            font-size: 1.3rem;
            font-weight: 800;
            letter-spacing: 1.2px;
            color: #7c3aed;
            text-transform: uppercase;
            margin-bottom: 1.5rem;
            margin-top: 0.2rem;
            padding-bottom: 0.8rem;
            border-bottom: 2px solid rgba(124,58,237,0.2);
        ">Features</div>
        """, unsafe_allow_html=True)
        
        # Custom styled navigation tabs
        st.markdown("""
        <style>
            .nav-tabs-container {
                display: flex;
                gap: 0.5rem;
                margin-bottom: 1.5rem;
                background: linear-gradient(135deg, rgba(255,255,255,0.8), rgba(248,250,252,0.6));
                padding: 0.6rem;
                border-radius: 14px;
                border: 1px solid rgba(124,58,237,0.15);
                box-shadow: 0 2px 8px rgba(0,0,0,0.04);
            }
            .nav-tab {
                flex: 1;
                padding: 0.65rem 0.75rem;
                border-radius: 10px;
                text-align: center;
                cursor: pointer;
                font-weight: 500;
                font-size: 0.85rem;
                transition: all 0.25s ease;
                border: 1.5px solid transparent;
                background: rgba(255,255,255,0.5);
                color: #64748b;
            }
            .nav-tab:hover {
                background: rgba(124,58,237,0.08);
                color: #7c3aed;
                border-color: rgba(124,58,237,0.2);
            }
            .nav-tab.active {
                background: linear-gradient(135deg, #7c3aed 0%, #6366f1 100%);
                color: white;
                border-color: #7c3aed;
                box-shadow: 0 4px 12px rgba(124,58,237,0.35);
            }
        </style>
        """, unsafe_allow_html=True)
        
        # Create custom navigation tabs (stacked vertically)
        current_nav = st.session_state.get("nav_page", "Workflow")
        
        if st.button("Workflow", use_container_width=True, key="btn_workflow"):
            st.session_state["nav_page"] = "Workflow"
        
        if st.button("History", use_container_width=True, key="btn_history"):
            st.session_state["nav_page"] = "History"
        
        if st.button("Feedback", use_container_width=True, key="btn_feedback"):
            st.session_state["nav_page"] = "Feedback"
        
        # Visual indicator of current tab
        st.markdown(f"""
        <div style="
            text-align: center;
            font-size: 0.8rem;
            color: #64748b;
            margin-top: 0.5rem;
            font-weight: 600;
            letter-spacing: 0.5px;
        ">→ {current_nav}</div>
        """, unsafe_allow_html=True)
        
        st.divider()
        # Show detected business domain, if available
        domain_info = st.session_state.get("document_domain_info")
        if domain_info:
            st.markdown("**🏷️ Detected domain**")
            st.caption(f"{domain_info.get('domain', 'General')} · confidence {domain_info.get('confidence', 0):.0%}")
            st.caption(domain_info.get("reason", ""))
            st.divider()

        st.markdown("""
        <div style="
            background: linear-gradient(135deg, rgba(13,148,136,0.12), rgba(124,58,237,0.1));
            border-radius: 12px; padding: 0.9rem 1rem; margin-bottom: 0.5rem;
            border: 1px solid rgba(124,58,237,0.15);
        ">
            <strong style="color: #1e293b;">🛡️ ReguGuard AI</strong><br>
            <span style="font-size: 0.8rem; color: #64748b;">Compliance analysis using latest regulations with AI.</span>
        </div>
        """, unsafe_allow_html=True)
        st.divider()
        st.markdown("**Features**")
        st.markdown("""
        ⭐ Pre-Audit Compliance Check \n
        ⭐ SOP Quality Improvement \n
        ⭐ Automated Documentation Review \n
        ⭐ Regulatory Gap Analysis Tool \n
        ⭐ Self-Learning Compliance Engine \n
        """)
        st.divider()
        st.markdown("**🌐 Regulation sources**")

        domain = st.session_state.get("document_domain")
        if domain:
            st.caption(f"*Domain-specific sources for: {domain}*")
            sources = get_regulation_sources(domain=domain, domain_only=True)
        else:
            sources = get_regulation_sources(domain=None)
        for src in sources[:8]:
            st.caption(f"· {src['name']}")
        if len(sources) > 8:
            st.caption(f"· +{len(sources)-8} more")
        st.divider()
        st.caption("Re-scrape regulations and re-run analysis to catch new gaps.")
        if st.button("Run monitoring check", use_container_width=True):
            st.session_state["current_step"] = "scrape"
            st.session_state["scraping_complete"] = False
            st.session_state["analysis_complete"] = False
            st.rerun()


def status_badge(status):
    badges = {
        "compliant": '<span class="rg-badge compliant">Compliant</span>',
        "partially_compliant": '<span class="rg-badge partial">Partial</span>',
        "non_compliant": '<span class="rg-badge non-compliant">Non-compliant</span>',
        "outdated": '<span class="rg-badge outdated">Outdated</span>',
    }
    return badges.get(status, status)


def risk_badge(risk):
    badges = {
        "High": '<span class="rg-risk-high">High</span>',
        "Medium": '<span class="rg-risk-medium">Medium</span>',
        "Low": '<span class="rg-risk-low">Low</span>',
    }
    return badges.get(risk, risk)


def render_upload_section():
    st.markdown('<p class="rg-section-title">📄 Upload document</p>', unsafe_allow_html=True)
    st.markdown('<p class="rg-section-caption">Upload your company SOP, policy, or compliance document (PDF, DOCX, or TXT). The system will read the full document and check if it is an SOP before running compliance analysis.</p>', unsafe_allow_html=True)

    col1, col2 = st.columns([2, 1])
    with col1:
        uploaded_file = st.file_uploader(
            "Choose file",
            type=["pdf", "docx", "txt"],
            key="sop_uploader",
            label_visibility="collapsed",
        )
        if uploaded_file is not None:
            st.caption(f"**{uploaded_file.name}** · {uploaded_file.size / 1024:.1f} KB")
            if st.button("Continue to document check", type="primary", key="parse_btn"):
                with st.spinner("Extracting text..."):
                    try:
                        file_bytes = uploaded_file.getvalue()
                        raw_text, pages_text = extract_text_and_pages_from_file(uploaded_file)
                        if not raw_text or not raw_text.strip():
                            st.error("No text could be extracted. Try another file.")
                            return
                        st.session_state["sop_text"] = raw_text
                        st.session_state["sop_pages"] = pages_text
                        st.session_state["uploaded_file_bytes"] = file_bytes
                        st.session_state["uploaded_file_name"] = uploaded_file.name
                        st.session_state["is_sop_result"] = None
                        # Reset domain detection for new document
                        st.session_state["document_domain_info"] = None
                        st.session_state["document_domain"] = None
                        st.session_state["current_step"] = "check_document"
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))
    with col2:
        st.caption("**Formats** · PDF, DOCX, TXT")
        if st.button("Try a sample SOP", key="sample_btn"):
            sample_text = _get_sample_sop()
            st.session_state["sop_text"] = sample_text
            st.session_state["is_sop_result"] = None
            # Reset domain detection for new document
            st.session_state["document_domain_info"] = None
            st.session_state["document_domain"] = None
            st.session_state["current_step"] = "check_document"
            st.rerun()


def render_check_document():
    st.markdown('<p class="rg-section-title">🔍 Document check</p>', unsafe_allow_html=True)
    sop_text = st.session_state.get("sop_text")
    if not sop_text:
        st.warning("No document in memory. Please upload a file first.")
        if st.button("Back to upload", key="back_upload_check"):
            st.session_state["current_step"] = "upload"
            st.rerun()
        return

    is_sop_result = st.session_state.get("is_sop_result")
    if is_sop_result is None:
        st.caption("Reading your document and checking if it is a company SOP or policy document...")
        with st.spinner("Analyzing document..."):
            use_ai = get_grok_client() is not None
            result = is_document_sop(sop_text, use_ai=use_ai)
            st.session_state["is_sop_result"] = result
        st.rerun()

    is_sop = is_sop_result.get("is_sop", False)
    reason = is_sop_result.get("reason", "")
    document_type = is_sop_result.get("document_type", "")

    if not is_sop:
        st.error("This does not appear to be a company SOP or policy document.")
        st.markdown(f'<div class="rg-card"><p><strong>Feedback:</strong> {reason}</p><p><strong>Detected type:</strong> {document_type}</p></div>', unsafe_allow_html=True)
        st.markdown('<p class="rg-section-caption">Upload a Standard Operating Procedure, company policy, compliance manual, or similar document to run compliance analysis.</p>', unsafe_allow_html=True)
        if st.button("Upload different file", type="primary", key="upload_different"):
            st.session_state["is_sop_result"] = None
            st.session_state["sop_text"] = None
            st.session_state["current_step"] = "upload"
            st.rerun()
        return

    # Detect business domain for domain-specific scraping
    if st.session_state.get("document_domain_info") is None:
        with st.spinner("Detecting document domain..."):
            domain_info = detect_document_domain(sop_text, use_ai=(get_grok_client() is not None))
            st.session_state["document_domain_info"] = domain_info
            st.session_state["document_domain"] = domain_info.get("domain")

    domain_info = st.session_state.get("document_domain_info") or {}
    domain_label = domain_info.get("domain", "General")
    domain_conf = domain_info.get("confidence", 0.0)

    st.success("Document recognized as a company SOP or policy document. You can proceed to fetch regulations and run compliance analysis.")
    st.markdown(
        f'<div class="rg-card"><p><strong>Reason:</strong> {reason}</p>'
        f'<p><strong>Detected domain:</strong> {domain_label} (confidence {domain_conf:.0%})</p></div>',
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Back to upload", key="back_upload_check"):
            st.session_state["current_step"] = "upload"
            st.rerun()
    with col2:
        if st.button("Fetch latest regulations", type="primary", key="proceed_scrape"):
            st.session_state["current_step"] = "scrape"
            st.rerun()


def render_scraping():
    st.markdown('<p class="rg-section-title">🌐 Fetch regulations</p>', unsafe_allow_html=True)
    st.markdown('<p class="rg-section-caption">Pulling current rules from government and institutional sources.</p>', unsafe_allow_html=True)

    if not st.session_state["scraping_complete"]:
        progress_bar = st.progress(0)
        status_text = st.empty()
        def progress_cb(current, total, name):
            progress_bar.progress((current + 1) / total)
            status_text.caption(f"Fetching: {name} ({current + 1}/{total})")
        with st.spinner("Loading..."):
            domain = st.session_state.get("document_domain")
            # Dynamic scraping: use ONLY domain-specific sources if domain detected
            if domain and domain != "General":
                sources = get_regulation_sources(domain=domain, domain_only=True)
                st.info(f"🔍 Scraping domain-specific sources for: **{domain}**")
            else:
                sources = get_regulation_sources(domain=None)
            regulations = scrape_regulations(sources=sources, progress_callback=progress_cb)
            st.session_state["regulations"] = regulations
            st.session_state["scraping_complete"] = True
            rag = RAGEngine()
            rag.index_regulations(regulations)
            st.session_state["rag_engine"] = rag
        progress_bar.progress(1.0)
        status_text.caption("Done.")
        st.rerun()
    else:
        regulations = st.session_state["regulations"]
        n_sources = len(set(r["source_name"] for r in regulations))
        st.caption(f"{len(regulations)} regulation entries from {n_sources} sources.")

        cats = {}
        for r in regulations:
            cat = r.get("category", "Unknown")
            cats[cat] = cats.get(cat, 0) + 1

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f'<div class="rg-metric accent-teal"><div class="value">{len(regulations)}</div><div class="label">Entries</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="rg-metric accent-violet"><div class="value">{len(cats)}</div><div class="label">Categories</div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="rg-metric accent-amber"><div class="value">{n_sources}</div><div class="label">Sources</div></div>', unsafe_allow_html=True)

        with st.expander("By category"):
            for cat, count in sorted(cats.items()):
                st.caption(f"{cat}: {count}")

        if st.button("Run compliance analysis", type="primary", key="proceed_analysis"):
            st.session_state["current_step"] = "analyze"
            st.rerun()


def render_analysis():
    st.markdown('<p class="rg-section-title">🔍 Compliance analysis</p>', unsafe_allow_html=True)

    sop_text = st.session_state.get("sop_text")
    rag = st.session_state.get("rag_engine")

    if not st.session_state["analysis_complete"]:
        if not sop_text:
            st.error("No document. Go back and upload an SOP.")
            return
        if rag is None:
            st.error("Regulations not loaded. Run regulation fetch first.")
            return

        st.caption("Reading your full document and comparing it against the latest regulations. This may take a moment.")
        progress_bar = st.progress(0)
        status_text = st.empty()
        regulations = st.session_state.get("regulations", [])

        status_text.caption("Analyzing full document...")
        progress_bar.progress(0.3)
        use_ai = get_grok_client() is not None
        analysis = analyze_full_document_compliance(sop_text, regulations, rag, use_ai=use_ai)
        progress_bar.progress(1.0)
        status_text.caption("Done.")

        st.session_state["full_document_analysis"] = analysis
        st.session_state["analysis_complete"] = True

        # Populate findings with best-effort location for history persistence
        sop_pages = st.session_state.get("sop_pages")
        for f in analysis.get("findings", []):
            if f.get("page") is None or f.get("line") is None or not f.get("text"):
                p, l, t = _find_doc_location(f.get("area"), f.get("issue"), f.get("suggestion"), sop_text, sop_pages)
                if f.get("page") is None: f["page"] = p
                if f.get("line") is None: f["line"] = l
                # Use found text if finding has no text or explicitly needs updating
                if not f.get("text") or f.get("text") == "N/A": 
                    f["text"] = t

        # Save to history
        try:
            document_name = st.session_state.get("uploaded_file_name", "Sample SOP")
            domain = st.session_state.get("document_domain")
            score = analysis.get("overall_score", 0)
            risk = analysis.get("overall_risk_level", "Medium")
            method = analysis.get("analysis_method", "Analysis")
            num_findings = len(analysis.get("findings", []))
            sop_preview = sop_text[:200] if sop_text else ""
            
            save_analysis_history(
                document_name=document_name,
                domain=domain,
                compliance_score=score,
                risk_level=risk,
                analysis_method=method,
                num_findings=num_findings,
                sop_text_preview=sop_preview,
                full_analysis=analysis,
                sop_text=sop_text,
                regulations=regulations,
            )
        except Exception as e:
            st.caption(f"Note: Could not save to history: {e}")
        
        st.session_state["current_step"] = "results"
        st.rerun()
    else:
        st.session_state["current_step"] = "results"
        st.rerun()


def render_results():
    st.markdown('<p class="rg-section-title">📊 Results</p>', unsafe_allow_html=True)

    analysis = st.session_state.get("full_document_analysis")

    if not analysis:
        st.warning("No results. Run analysis first.")
        if st.button("Go to upload", key="go_upload"):
            st.session_state["current_step"] = "upload"
            st.rerun()
        return

    score = analysis.get("overall_score", 0)
    risk_level = analysis.get("overall_risk_level", "Medium")
    score_class = "score-good" if score >= 70 else "score-mid" if score >= 40 else "score-low"
    findings = analysis.get("findings", [])
    debt = analysis.get("compliance_debt_estimate", 0)
    method = analysis.get("analysis_method", "Analysis")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div class="rg-metric rg-hero-metric {score_class} accent-teal"><div class="value">{score}%</div><div class="label">Compliance Score</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="rg-metric accent-amber"><div class="value">{risk_badge(risk_level)}</div><div class="label">Overall Risk</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="rg-metric accent-violet"><div class="value">${debt:,.0f}</div><div class="label">Est. compliance debt</div></div>', unsafe_allow_html=True)

    st.markdown("<div class='rg-divider'></div>", unsafe_allow_html=True)

    executive_summary = analysis.get("executive_summary", "")
    if executive_summary:
        st.markdown("**Executive summary**")
        st.markdown(f'<div class="rg-card"><p>{executive_summary}</p></div>', unsafe_allow_html=True)
        st.markdown("<div class='rg-divider'></div>", unsafe_allow_html=True)

    st.markdown('<p class="rg-section-title">📑 Findings (from your document)</p>', unsafe_allow_html=True)
    st.caption("Feedback is generated dynamically based on the content you uploaded.")

    filter_status = st.selectbox(
        "Filter by status",
        ["All", "Compliant", "Partially Compliant", "Non-Compliant", "Outdated"],
        key="filter_status",
        label_visibility="collapsed",
    )
    status_map = {"Compliant": "compliant", "Partially Compliant": "partially_compliant", "Non-Compliant": "non_compliant", "Outdated": "outdated"}

    sort_risk = st.selectbox(
        "Sort by risk",
        ["None", "High → Low", "Low → High"],
        key="sort_risk",
        label_visibility="collapsed",
    )

    risk_rank = {"High": 3, "Medium": 2, "Low": 1}
    findings_with_idx = list(enumerate(findings))
    if sort_risk == "High → Low":
        findings_with_idx.sort(key=lambda it: risk_rank.get(it[1].get("risk_level", "Medium"), 2), reverse=True)
    elif sort_risk == "Low → High":
        findings_with_idx.sort(key=lambda it: risk_rank.get(it[1].get("risk_level", "Medium"), 2))

    for i, f in findings_with_idx:
        status = f.get("status", "partially_compliant")
        if filter_status != "All" and status != status_map.get(filter_status):
            continue
        area = f.get("area", "Section")
        risk = f.get("risk_level", "Medium")
        issue = f.get("issue", "")
        suggestion = f.get("suggestion", "")
        applicable = f.get("applicable_regulations", [])

        with st.expander(f"{area} — {status.replace('_', ' ').title()} · Risk: {risk}", expanded=(status != "compliant")):
            st.markdown(f"**Status:** {status_badge(status)}  **Risk:** {risk_badge(risk)}", unsafe_allow_html=True)
            if issue and str(issue).lower() != "none":
                st.markdown(f"**Issue:** {issue}")

            if suggestion and str(suggestion).lower() not in ("no change needed", "none"):
                st.markdown("**Suggestion**")
                st.markdown(f'<div class="rg-suggestion">{suggestion}</div>', unsafe_allow_html=True)
                
                finding_key = f"{area}_{i}"
                col_approve, col_reject = st.columns(2)
                
                with col_approve:
                    # Marker div used only for CSS targeting (green approve button)
                    st.markdown('<div class="rg-approve-marker"></div>', unsafe_allow_html=True)
                    approve_key = f"approve_finding_{i}_{area[:30]}"
                    if st.button("Approve change", key=approve_key, type="primary"):
                        st.session_state.setdefault("approved_changes", {})[finding_key] = {
                            "suggestion": suggestion,
                            "approved_at": datetime.now().isoformat(),
                            "area": area,
                            "issue": issue,
                            "status": status,
                            "risk_level": risk,
                        }
                        # Remove from rejected if it was there
                        if finding_key in st.session_state.get("rejected_changes", {}):
                            del st.session_state["rejected_changes"][finding_key]
                        st.success("Approved.")
                        st.rerun()
                
                with col_reject:
                    # Marker div used only for CSS targeting (red reject button)
                    st.markdown('<div class="rg-reject-marker"></div>', unsafe_allow_html=True)
                    reject_key = f"reject_finding_{i}_{area[:30]}"
                    if st.button("Reject change", key=reject_key):
                        st.session_state.setdefault("rejected_changes", {})[finding_key] = {
                            "suggestion": suggestion,
                            "rejected_at": datetime.now().isoformat(),
                            "area": area,
                            "issue": issue,
                            "status": status,
                            "risk_level": risk,
                        }
                        # Remove from approved if it was there
                        if finding_key in st.session_state.get("approved_changes", {}):
                            del st.session_state["approved_changes"][finding_key]
                        st.info("Rejected.")
                        st.rerun()
                
                # Show current status
                if st.session_state.get("approved_changes", {}).get(finding_key):
                    st.caption(f"✅ Approved · {st.session_state['approved_changes'][finding_key]['approved_at'][:10]}")
                elif st.session_state.get("rejected_changes", {}).get(finding_key):
                    st.caption(f"❌ Rejected · {st.session_state['rejected_changes'][finding_key]['rejected_at'][:10]}")
            if applicable:
                st.markdown("**Applicable regulations:**")
                for reg in applicable[:5]:
                    st.caption(f"· {reg}")

    missing = analysis.get("missing_elements", [])
    if missing:
        st.markdown("<div class='rg-divider'></div>", unsafe_allow_html=True)
        st.markdown("**Missing regulatory elements**")
        for m in missing:
            st.markdown(f"- {m}")

    citations = analysis.get("citations", [])
    if citations:
        st.markdown("<div class='rg-divider'></div>", unsafe_allow_html=True)
        st.markdown("**Citations & sources**")
        st.caption("Regulatory sources used for this analysis: government and institutional pages that were matched to your document. Each link points to the official regulation or guidance (e.g. OSHA, EPA, HIPAA).")
        st.caption("**Relevance %** = how closely that regulation’s text matches your document (0–100%). Higher means more overlap in topics and wording; it is computed automatically from the document and regulation text.")
        for cit in citations:
            url = cit.get("url", "#")
            rev_pct = cit.get("relevance_score", 0)
            st.markdown(f"""<div class="rg-citation">
                <strong>{cit.get('source', 'Unknown')}</strong> — {cit.get('title', 'N/A')}<br>
                <a href="{url}" target="_blank" rel="noopener">Source</a> · Relevance {rev_pct:.0%}
            </div>""", unsafe_allow_html=True)

    st.markdown("<div class='rg-divider'></div>", unsafe_allow_html=True)
    approved = st.session_state.get("approved_changes", {})
    if approved:
        st.markdown('<p class="rg-section-title">✅ Approved changes</p>', unsafe_allow_html=True)
        st.markdown(f'<p class="rg-section-caption">{len(approved)} suggestion(s) approved — added to the updated document in real time. The updated SOP contains the full original text plus each approved regulatory update inserted after the relevant section, with clear labels and spacing.</p>', unsafe_allow_html=True)
        for key, change in approved.items():
            st.caption(f"· {change.get('area', key)} — {change.get('approved_at', '')[:10]}")


    st.markdown("<div class='rg-divider'></div>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        try:
            excel_data = _generate_excel_export(analysis)
            st.download_button(
                "📊 Download Excel Sheet",
                data=excel_data,
                file_name=f"compliance_issues_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel",
            )
        except Exception as e:
            st.error(f"Error generating Excel: {str(e)}")
    with col2:
        if st.button("Run analysis again", key="rerun_analysis"):
            st.session_state["analysis_complete"] = False
            st.session_state["current_step"] = "analyze"
            st.rerun()
    with col3:
        if st.button("Start a new analysis", key="new_analysis"):
            for key in ["sop_text", "uploaded_file_bytes", "uploaded_file_name", "is_sop_result", "regulations", "rag_engine", "full_document_analysis",
                        "approved_changes", "rejected_changes", "analysis_complete", "scraping_complete"]:
                if key in st.session_state:
                    del st.session_state[key]
            st.session_state["current_step"] = "upload"
            st.rerun()


def render_history():
    """Render the History tab showing past analyses (ChatGPT-style)."""
    # Analysis history is displayed as a list below
    
    history = load_analysis_history(limit=100)
    
    if not history:
        st.info("No analysis history yet. Complete an analysis to see it here.")
        if st.button("Go to Workflow", type="primary"):
            st.session_state["current_step"] = "upload"
            st.rerun()
        return
    
    # Show selected analysis or list all
    selected_id = st.session_state.get("selected_history_id")
    
    if selected_id:
        # Show full analysis details
        record = get_analysis_by_id(selected_id)
        if record:
            # Header with analysis info
            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"### {record.get('document_name', 'Unknown Document')}")
                st.caption(f"Domain: {record.get('domain', 'General')} · {record.get('analysis_method', 'Analysis')}")
            with col2:
                col_back, col_del = st.columns(2)
                with col_back:
                    if st.button("← Back", key="back_to_history_list"):
                        st.session_state["selected_history_id"] = None
                        st.rerun()
                with col_del:
                    if st.button("🗑️ Delete", key=f"delete_detail_{selected_id}"):
                        if delete_analysis(selected_id):
                            st.success("Analysis deleted! Returning to list...")
                            st.session_state["selected_history_id"] = None
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error("Could not delete analysis")
            
            st.divider()
            
            # Display metrics
            score = record.get("compliance_score", 0)
            risk_level = record.get("risk_level", "Medium")
            score_class = "score-good" if score >= 70 else "score-mid" if score >= 40 else "score-low"
            
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f'<div class="rg-metric rg-hero-metric {score_class} accent-teal"><div class="value">{score}%</div><div class="label">Compliance Score</div></div>', unsafe_allow_html=True)
            with c2:
                st.markdown(f'<div class="rg-metric accent-amber"><div class="value">{risk_badge(risk_level)}</div><div class="label">Overall Risk</div></div>', unsafe_allow_html=True)
            with c3:
                st.markdown(f'<div class="rg-metric accent-violet"><div class="value">{record.get("num_findings", 0)}</div><div class="label">Findings</div></div>', unsafe_allow_html=True)
            
            st.divider()
            
            # Download buttons for Excel
            col_dlx1, col_dlx2 = st.columns(2)
            with col_dlx1:
                try:
                    excel_data = _generate_history_detailed_excel_export(record)
                    st.download_button(
                        label="📊 Download Detailed Excel",
                        data=excel_data,
                        file_name=f"analysis_{record.get('analysis_id', 'unknown')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_detailed_excel_{selected_id}",
                    )
                except Exception as e:
                    st.warning(f"Could not generate Excel: {str(e)[:100]}")
            
            with col_dlx2:
                try:
                    excel_data = _generate_history_excel_export(record)
                    st.download_button(
                        label="📊 Download Summary Excel",
                        data=excel_data,
                        file_name=f"summary_{record.get('analysis_id', 'unknown')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_summary_excel_{selected_id}",
                    )
                    st.caption("Summary, Findings & Sources sheets")
                except Exception as e:
                    st.warning(f"Could not generate summary: {str(e)[:100]}")
            
            st.divider()
            
            # Show timestamp
            timestamp = record.get("timestamp", "")
            try:
                dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
                time_str = dt.strftime("%B %d, %Y at %I:%M %p")
            except:
                time_str = timestamp
            st.caption(f"Analysis completed: {time_str}")
            
            st.divider()
            
            # Show full analysis if available
            full_analysis = record.get("full_analysis")
            if full_analysis:
                # Executive summary
                exec_summary = full_analysis.get("executive_summary", "")
                if exec_summary:
                    st.markdown("**Executive Summary**")
                    st.markdown(f'<div class="rg-card"><p>{exec_summary}</p></div>', unsafe_allow_html=True)
                    st.divider()
                
                # Findings
                findings = full_analysis.get("findings", [])
                if findings:
                    st.markdown("**Findings**")
                    for i, f in enumerate(findings):
                        area = f.get("area", "Section")
                        status = f.get("status", "partially_compliant")
                        risk = f.get("risk_level", "Medium")
                        issue = f.get("issue", "")
                        suggestion = f.get("suggestion", "")
                        applicable = f.get("applicable_regulations", [])
                        
                        with st.expander(f"{area} — {status.replace('_', ' ').title()} · Risk: {risk}", expanded=(status != "compliant")):
                            st.markdown(f"**Status:** {status_badge(status)}  **Risk:** {risk_badge(risk)}", unsafe_allow_html=True)
                            if issue and str(issue).lower() != "none":
                                st.markdown(f"**Issue:** {issue}")
                            if suggestion and str(suggestion).lower() not in ("no change needed", "none"):
                                st.markdown("**Suggestion**")
                                st.markdown(f'<div class="rg-suggestion">{suggestion}</div>', unsafe_allow_html=True)
                            if applicable:
                                st.markdown("**Applicable regulations:**")
                                for reg in applicable[:5]:
                                    st.caption(f"· {reg}")
                
                # Missing elements
                missing = full_analysis.get("missing_elements", [])
                if missing:
                    st.divider()
                    st.markdown("**Missing Regulatory Elements**")
                    for m in missing:
                        st.markdown(f"- {m}")
                
                # Citations
                citations = full_analysis.get("citations", [])
                if citations:
                    st.divider()
                    st.markdown("**Citations & Sources**")
                    st.caption("Regulatory sources used for this analysis: government and institutional pages that were matched to your document.")
                    for cit in citations[:15]:
                        url = cit.get("url", "#")
                        rev_pct = cit.get("relevance_score", 0)
                        st.markdown(f"""<div class="rg-citation">
                            <strong>{cit.get('source', 'Unknown')}</strong> — {cit.get('title', 'N/A')}<br>
                            <a href="{url}" target="_blank" rel="noopener">Source</a> · Relevance {rev_pct:.0%}
                        </div>""", unsafe_allow_html=True)
            else:
                # Show summary even if full_analysis is missing (for older records)
                st.markdown("**Analysis Summary**")
                st.markdown(f'<div class="rg-card">', unsafe_allow_html=True)
                st.markdown(f"""
                <p><strong>Compliance Score:</strong> {score}%</p>
                <p><strong>Risk Level:</strong> {risk_level}</p>
                <p><strong>Number of Findings:</strong> {record.get('num_findings', 0)}</p>
                <p><strong>Analysis Method:</strong> {record.get('analysis_method', 'Unknown')}</p>
                <p><strong>Domain:</strong> {record.get('domain', 'General')}</p>
                </div>
                """, unsafe_allow_html=True)
                st.info("⚠️ Full detailed findings are not available for this older record. New analyses will include complete details.")
            
            # Show SOP preview if available
            sop_preview = record.get("sop_text_preview") or record.get("sop_text")
            if sop_preview:
                st.divider()
                with st.expander("Document Preview"):
                    st.text(sop_preview[:1000] + ("..." if len(sop_preview) > 1000 else ""))
        else:
            st.error("Analysis not found.")
            if st.button("Back to list"):
                st.session_state["selected_history_id"] = None
                st.rerun()
    else:
        st.divider()
        
        for record in history:
            score = record.get("compliance_score", 0)
            risk = record.get("risk_level", "Medium")
            doc_name = record.get("document_name", "Unknown")
            timestamp = record.get("timestamp", "")
            domain = record.get("domain", "General")
            analysis_id = record.get("analysis_id", "")
            num_findings = record.get("num_findings", 0)
            
            # Format timestamp
            try:
                dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
                time_str = dt.strftime("%b %d, %Y")
            except:
                time_str = timestamp[:10] if timestamp else "Unknown"
            
            # Count findings breakdown
            full_analysis = record.get("full_analysis") or {}
            findings = full_analysis.get("findings", [])
            compliant_count = sum(1 for f in findings if f.get("status") == "compliant")
            partial_count = sum(1 for f in findings if f.get("status") == "partially_compliant")
            non_compliant_count = sum(1 for f in findings if f.get("status") == "non_compliant")
            
            # Get brief summary
            exec_summary = full_analysis.get("executive_summary", "")
            brief_summary = exec_summary[:150] + "..." if len(exec_summary) > 150 else exec_summary
            
            # Show analysis card
            with st.container():
                st.markdown(
                    f"""
                    <div style="background: rgba(255,255,255,0.7); border: 1px solid rgba(0,0,0,0.1); border-radius: 10px; padding: 1rem; margin-bottom: 1rem;">
                        <div style="display: flex; justify-content: space-between; align-items: start;">
                            <div style="flex: 1;">
                                <h4 style="margin: 0; color: #1e293b;">{doc_name}</h4>
                                <p style="margin: 0.3rem 0; font-size: 0.9rem; color: #64748b;">{domain} · {time_str}</p>
                                <p style="margin: 0.5rem 0; font-size: 0.85rem; color: #475569; line-height: 1.4;">{brief_summary if brief_summary else "No summary available"}</p>
                                <div style="display: flex; gap: 1rem; margin-top: 0.5rem; font-size: 0.8rem;">
                                    <span style="color: #10b981;">✓ Compliant: {compliant_count}</span>
                                    <span style="color: #f59e0b;">⊘ Partial: {partial_count}</span>
                                    <span style="color: #ef4444;">✗ Non-Compliant: {non_compliant_count}</span>
                                </div>
                            </div>
                            <div style="text-align: center; margin-left: 1rem;">
                                <div style="background: {'#10b981' if score >= 70 else '#f59e0b' if score >= 40 else '#ef4444'}; color: white; padding: 0.7rem 1rem; border-radius: 8px; font-weight: bold; font-size: 1.1rem;">
                                    {score}%
                                </div>
                                <p style="margin: 0.3rem 0; font-size: 0.75rem; color: #64748b;">{risk} Risk</p>
                            </div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                
                # Action buttons
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    if st.button("👁️ View", key=f"view_{analysis_id}", use_container_width=True):
                        st.session_state["selected_history_id"] = analysis_id
                        st.rerun()
                
                with col2:
                    try:
                        excel_data = _generate_history_detailed_excel_export(record)
                        st.download_button(
                            label="📥 Excel",
                            data=excel_data,
                            file_name=f"analysis_{analysis_id}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_excel_{analysis_id}",
                            use_container_width=True,
                        )
                    except Exception as e:
                        st.button("📥 Excel", disabled=True, use_container_width=True)
                
                with col3:
                    if st.button("📋 Summary", key=f"summary_{analysis_id}", use_container_width=True):
                        st.session_state[f"show_summary_{analysis_id}"] = not st.session_state.get(f"show_summary_{analysis_id}", False)
                        st.rerun()
                
                with col4:
                    if st.button("🗑️ Delete", key=f"delete_{analysis_id}", use_container_width=True):
                        if delete_analysis(analysis_id):
                            st.success(f"Analysis deleted!")
                            st.rerun()
                        else:
                            st.error("Could not delete analysis")
                
                # Show summary if toggled
                if st.session_state.get(f"show_summary_{analysis_id}", False):
                    st.markdown(
                        f"""
                        <div style="background: #f0f9ff; border-left: 4px solid #0ea5e9; padding: 1rem; border-radius: 4px; margin: 0.5rem 0;">
                            <strong style="color: #0369a1;">Summary</strong><br>
                            <span style="font-size: 0.9rem; color: #1e293b;">{brief_summary if brief_summary else "No summary available"}</span>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                
                st.divider()


def render_feedback():
    import difflib
    import hashlib

    st.markdown('<p class="rg-section-title">🗣️ Feedback</p>', unsafe_allow_html=True)
    st.markdown(
        '<p class="rg-section-caption">Upload the ReguGuard-updated SOP and the human-updated SOP. We will compare them and store a feedback record you can use to improve future accuracy.</p>',
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**AI-updated SOP (ReguGuard output)**")
        use_last = st.button("Use last generated updated SOP", use_container_width=True)
        ai_file = st.file_uploader("AI Updated SOP", type=["pdf", "docx", "txt"], key="fb_ai_file", label_visibility="collapsed")
    with c2:
        st.markdown("**Human-updated SOP**")
        human_file = st.file_uploader("Human Updated SOP", type=["pdf", "docx", "txt"], key="fb_human_file", label_visibility="collapsed")

    ai_name = None
    ai_bytes = None
    if use_last and st.session_state.get("last_updated_sop_bytes") and st.session_state.get("last_updated_sop_name"):
        ai_name = st.session_state.get("last_updated_sop_name")
        ai_bytes = st.session_state.get("last_updated_sop_bytes")
        st.success(f"Loaded last generated file: {ai_name}")
    elif ai_file is not None:
        ai_name = ai_file.name
        ai_bytes = ai_file.getvalue()

    human_name = None
    human_bytes = None
    if human_file is not None:
        human_name = human_file.name
        human_bytes = human_file.getvalue()

    if not ai_bytes or not human_bytes:
        st.info("Upload both files (or use the last generated updated SOP) to compare.")
        return

    try:
        ai_text = extract_text_from_bytes(ai_name, ai_bytes)
        human_text = extract_text_from_bytes(human_name, human_bytes)
    except Exception as e:
        st.error(f"Could not read one of the files: {e}")
        return

    ai_lines = (ai_text or "").splitlines()
    human_lines = (human_text or "").splitlines()

    sm = difflib.SequenceMatcher(None, ai_text or "", human_text or "")
    similarity = sm.ratio()
    diff = list(difflib.unified_diff(ai_lines, human_lines, fromfile=ai_name or "ai", tofile=human_name or "human", lineterm=""))

    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("Similarity", f"{similarity*100:.1f}%")
    with m2:
        st.metric("AI lines", str(len(ai_lines)))
    with m3:
        st.metric("Human lines", str(len(human_lines)))

    with st.expander("View differences (AI vs Human)", expanded=True):
        st.text("\n".join(diff[:2000]) + ("\n...\n(truncated)" if len(diff) > 2000 else ""))

    # Persist feedback record for future improvement (dataset collection)
    feedback_dir = os.path.join("data", "feedback")
    os.makedirs(feedback_dir, exist_ok=True)
    record = {
        "created_at": datetime.now().isoformat(),
        "ai_filename": ai_name,
        "human_filename": human_name,
        "ai_sha256": hashlib.sha256(ai_bytes).hexdigest(),
        "human_sha256": hashlib.sha256(human_bytes).hexdigest(),
        "similarity": similarity,
        "diff_unified": "\n".join(diff[:5000]),
    }
    record_path = os.path.join(feedback_dir, "feedback_log.jsonl")
    try:
        with open(record_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
        st.caption(f"Feedback saved to `{record_path}` (append-only).")
    except Exception as e:
        st.warning(f"Could not save feedback log: {e}")

    st.download_button(
        "Download diff report (txt)",
        data="\n".join(diff),
        file_name=f"reguguard_feedback_diff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
        mime="text/plain",
        use_container_width=True,
    )
    st.download_button(
        "Download feedback record (json)",
        data=json.dumps(record, indent=2),
        file_name=f"reguguard_feedback_record_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        mime="application/json",
        use_container_width=True,
    )


def _generate_report(analysis):
    return {
        "report_title": "ReguGuard AI Compliance Report",
        "generated_at": datetime.now().isoformat(),
        "document_analysis": analysis,
        "approved_changes": st.session_state.get("approved_changes", {}),
    }


def _pick_best_citation_for_finding(applicable_regs, citations):
    if not citations or not applicable_regs:
        return None
        
    for reg_str in applicable_regs:
        reg_str = str(reg_str or "").strip()
        if not reg_str:
            continue
            
        # 1. Try matching by ID if present in the string (e.g. "Regulation [1]")
        match = re.search(r"\[(\d+)\]", reg_str)
        if match:
            reg_id = int(match.group(1))
            for cit in citations:
                if cit.get("id") == reg_id:
                    return cit
                    
        # 2. Try matching by Name/Source similarity
        for cit in citations:
            title = str(cit.get("title", "") or "").lower()
            source = str(cit.get("source", "") or "").lower()
            reg_lower = reg_str.lower()
            if reg_lower in title or reg_lower in source or title in reg_lower or source in reg_lower:
                return cit
                
    # 3. Fallback to first citation only if very few citations exist (likely same source)
    if len(citations) == 1:
        return citations[0]
        
    return None


def _find_doc_location(area, issue, suggestion, sop_text, sop_pages):
    """
    Best-effort locator for where the finding occurs in the uploaded document.
    Returns: (page_no | None, line_no | None, matched_line_text | "")
    - For PDFs: line_no is the line number within the matched page (best effort).
    - For non-PDFs: line_no is the line number within the full text.
    """
    if not sop_text:
        return None, None, ""

    def _tokens(text):
        if not text:
            return []
        # keep useful words only
        raw = re.findall(r"[A-Za-z0-9][A-Za-z0-9\-_/]{2,}", str(text).lower())
        stop = {
            "the","and","for","with","that","this","from","into","your","you","are","was","were","will","shall",
            "should","must","may","can","a","an","to","of","in","on","at","as","by","or","is","it","be","not",
            "review","add","document","procedures","procedure","policy","section"
        }
        toks = [t for t in raw if t not in stop and len(t) >= 4]
        # de-dup while preserving order
        seen = set()
        out = []
        for t in toks:
            if t not in seen:
                seen.add(t)
                out.append(t)
        return out[:20]

    # Tokens from issue/suggestion/area to find the "replacement" line context
    toks = []
    toks += _tokens(issue)
    toks += _tokens(suggestion)
    toks += _tokens(area)
    # de-dup again
    seen = set()
    tokens = []
    for t in toks:
        if t not in seen:
            seen.add(t)
            tokens.append(t)
    tokens = tokens[:20]

    def _score_line(line):
        ll = (line or "").lower()
        if not ll.strip():
            return 0
        score = 0
        for t in tokens:
            if t in ll:
                score += 2
        # slight boost if whole issue fragment appears
        if issue and str(issue).strip() and str(issue).strip().lower() in ll:
            score += 5
        return score

    # If PDF pages are available, find best matching line within pages first
    if sop_pages:
        best = {"score": 0, "page": None, "line": None, "text": ""}
        for p_idx, p_text in enumerate(sop_pages, start=1):
            page_lines = (p_text or "").splitlines()
            for l_idx, line in enumerate(page_lines, start=1):
                s = _score_line(line)
                if s > best["score"]:
                    best = {"score": s, "page": p_idx, "line": l_idx, "text": (line or "").strip()}
        if best["score"] > 0 and best["text"]:
            return best["page"], best["line"], best["text"]

    # Fallback: score full document lines
    lines = sop_text.splitlines()
    best_idx = None
    best_score = 0
    for idx, line in enumerate(lines):
        s = _score_line(line)
        if s > best_score:
            best_score = s
            best_idx = idx

    if best_idx is not None and best_score > 0:
        return None, best_idx + 1, (lines[best_idx] or "").strip()

    # If nothing matched, return first non-empty line as last resort
    for idx, line in enumerate(lines):
        if (line or "").strip():
            return None, idx + 1, (line or "").strip()

    return None, None, ""


def _generate_excel_export(analysis):
    """
    Generate an Excel file with compliance issues, source page numbers, 
    approve/reject actions, and date column.
    """
    findings = analysis.get("findings", [])
    citations = analysis.get("citations", [])
    approved_changes = st.session_state.get("approved_changes", {})
    rejected_changes = st.session_state.get("rejected_changes", {})
    sop_text = st.session_state.get("sop_text")
    sop_pages = st.session_state.get("sop_pages")
    
    # Create a mapping of citations to page numbers (using citation index as page reference)
    citation_map = {}
    for idx, cit in enumerate(citations, start=1):
        citation_map[cit.get("source", "")] = idx
    
    # Build Excel data
    excel_data = []
    
    for i, finding in enumerate(findings):
        area = finding.get("area", "Section")
        issue = finding.get("issue", "")
        status = finding.get("status", "")
        risk = finding.get("risk_level", "") or "Medium"
        suggestion = finding.get("suggestion", "")
        applicable_regs = finding.get("applicable_regulations", [])
        
        # Get source link from applicable regulations or citations
        source_link = ""
        matched_cit = _pick_best_citation_for_finding(applicable_regs, citations)
        if matched_cit:
            source_link = str(matched_cit.get("url", "") or "")
            source_page_no = citations.index(matched_cit) + 1 if matched_cit in citations else 1
        else:
            source_page_no = 1
        
        # Find best-effort location in uploaded document
        doc_page, doc_line, doc_text = _find_doc_location(area, issue, suggestion, sop_text, sop_pages)

        finding_key = f"{area}_{i}"
        
        # Determine action
        action = "Pending"
        action_date = ""
        if finding_key in approved_changes:
            action = "Approve"
            action_date = approved_changes[finding_key].get("approved_at", "")[:10]
        elif finding_key in rejected_changes:
            action = "Reject"
            action_date = rejected_changes[finding_key].get("rejected_at", "")[:10]
        
        # Create compliance issue description
        error_suggestion = ""
        if issue and str(issue).lower() != "none":
            error_suggestion += f"Error: {issue}"
        if suggestion and str(suggestion).lower() not in ("no change needed", "none"):
            error_suggestion += (" | " if error_suggestion else "") + f"Suggestion: {suggestion}"
        if not error_suggestion:
            error_suggestion = "N/A"
        
        excel_data.append({
            "Risk Level": risk,
            "Page": doc_page if doc_page is not None else "N/A",
            "Line": doc_line if doc_line is not None else "N/A",
            "Text": doc_text if doc_text else "N/A",
            "Error / Suggestion": error_suggestion,
            "Source Link": source_link if source_link else "N/A",
            "Action": action,
            "Date": action_date if action_date else datetime.now().strftime("%Y-%m-%d"),
        })
    
    # Create DataFrame - ensure we have at least headers even if no data
    if not excel_data:
        excel_data = [{
            "Risk Level": "Low",
            "Page": "N/A",
            "Line": "N/A",
            "Text": "No compliance issues found",
            "Error / Suggestion": "N/A",
            "Source Link": "N/A",
            "Action": "Pending",
            "Date": datetime.now().strftime("%Y-%m-%d"),
        }]
    
    df = pd.DataFrame(excel_data)

    # Sort by risk severity (High -> Medium -> Low)
    risk_rank = {"High": 3, "Medium": 2, "Low": 1}
    df["_risk_rank"] = df["Risk Level"].map(lambda x: risk_rank.get(str(x), 2))
    df = df.sort_values(by="_risk_rank", ascending=False).drop(columns=["_risk_rank"])
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Compliance Issues')
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Compliance Issues']
        
        # Style the header row
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 12  # Risk Level
        worksheet.column_dimensions['B'].width = 8   # Page
        worksheet.column_dimensions['C'].width = 8   # Line
        worksheet.column_dimensions['D'].width = 55  # Text
        worksheet.column_dimensions['E'].width = 55  # Error / Suggestion
        worksheet.column_dimensions['F'].width = 55  # Source Link
        worksheet.column_dimensions['G'].width = 12  # Action
        worksheet.column_dimensions['H'].width = 15  # Date
        
        # Make Date column editable by ensuring it's formatted as date
        for row in range(2, len(df) + 2):
            date_cell = worksheet[f'H{row}']
            # Set as date format - users can edit this
            date_cell.number_format = 'YYYY-MM-DD'
            date_cell.alignment = Alignment(horizontal="center")
            
            # Also make Action column centered
            action_cell = worksheet[f'G{row}']
            action_cell.alignment = Alignment(horizontal="center")

            # Wrap long text fields
            worksheet[f'D{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            worksheet[f'E{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            worksheet[f'F{row}'].alignment = Alignment(wrap_text=True, vertical="top")
    
    output.seek(0)
    return output.getvalue()


def _generate_history_detailed_excel_export(record):
    """
    Generate a detailed Excel file from a history record, using exact same format as workflow tab.
    Uses the preserved page/line numbers from the original analysis.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    full_analysis = record.get("full_analysis") or {}
    findings = full_analysis.get("findings", [])
    citations = full_analysis.get("citations", [])
    sop_text = record.get("sop_text", "")
    
    # Build Excel data
    excel_data = []
    
    for i, finding in enumerate(findings):
        area = finding.get("area", "Section")
        issue = finding.get("issue", "")
        status = finding.get("status", "")
        risk = finding.get("risk_level", "") or "Medium"
        suggestion = finding.get("suggestion", "")
        applicable_regs = finding.get("applicable_regulations", [])
        
        # Use page/line from finding if available, otherwise use defaults
        doc_page = finding.get("page") if finding.get("page") is not None else "N/A"
        doc_line = finding.get("line") if finding.get("line") is not None else "N/A"
        doc_text = finding.get("text", "N/A") if finding.get("text") else "N/A"
        
        # Get source link from applicable regulations or citations
        source_link = ""
        matched_cit = _pick_best_citation_for_finding(applicable_regs, citations)
        if matched_cit:
            source_link = str(matched_cit.get("url", "") or "")
        
        # Create compliance issue description
        error_suggestion = ""
        if issue and str(issue).lower() != "none":
            error_suggestion += f"Error: {issue}"
        if suggestion and str(suggestion).lower() not in ("no change needed", "none"):
            error_suggestion += (" | " if error_suggestion else "") + f"Suggestion: {suggestion}"
        if not error_suggestion:
            error_suggestion = "N/A"
        
        # Get analysis date
        action_date = record.get("timestamp", "")[:10]
        
        excel_data.append({
            "Risk Level": risk,
            "Page": doc_page,
            "Line": doc_line,
            "Text": str(doc_text)[:200] if doc_text else "N/A",
            "Error / Suggestion": error_suggestion,
            "Source Link": source_link if source_link else "N/A",
            "Date": action_date if action_date else datetime.now().strftime("%Y-%m-%d"),
        })
    
    # Create DataFrame - ensure we have at least headers even if no data
    if not excel_data:
        excel_data = [{
            "Risk Level": "Low",
            "Page": "N/A",
            "Line": "N/A",
            "Text": "No compliance issues found",
            "Error / Suggestion": "N/A",
            "Source Link": "N/A",
            "Date": datetime.now().strftime("%Y-%m-%d"),
        }]
    
    df = pd.DataFrame(excel_data)

    # Sort by risk severity (High -> Medium -> Low)
    risk_rank = {"High": 3, "Medium": 2, "Low": 1}
    df["_risk_rank"] = df["Risk Level"].map(lambda x: risk_rank.get(str(x), 2))
    df = df.sort_values(by="_risk_rank", ascending=False).drop(columns=["_risk_rank"])
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Compliance Issues')
        
        # Get workbook and worksheet to format
        workbook = writer.book
        worksheet = writer.sheets['Compliance Issues']
        
        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths and add styling
        col_widths = {"A": 12, "B": 8, "C": 8, "D": 20, "E": 35, "F": 30, "G": 12}
        for col, width in col_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Format data rows
        for row in range(2, worksheet.max_row + 1):
            risk_cell = worksheet[f'A{row}']
            risk_value = str(risk_cell.value or "")
            
            # Color-code risk levels
            if "High" in risk_value:
                risk_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                risk_cell.font = Font(bold=True, color="FFFFFF")
            elif "Medium" in risk_value:
                risk_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                risk_cell.font = Font(bold=True, color="FFFFFF")
            elif "Low" in risk_value:
                risk_cell.fill = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")
                risk_cell.font = Font(bold=True, color="FFFFFF")
            
            # Wrap text in description columns
            worksheet[f'D{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            worksheet[f'E{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            worksheet[f'F{row}'].alignment = Alignment(wrap_text=True, vertical="top")
    
    output.seek(0)
    return output.getvalue()


def _generate_history_excel_export(record):
    """
    Generate an Excel file from a single analysis record for history download.
    Includes summary, findings, and citations sheets.
    """
    # Prepare data for Excel sheets
    summary_data = {
        "Parameter": [
            "Document Name",
            "Domain",
            "Analysis Date",
            "Compliance Score",
            "Risk Level",
            "Total Findings",
            "Analysis Method",
        ],
        "Value": [
            record.get("document_name", "Unknown"),
            record.get("domain", "General"),
            record.get("timestamp", "")[:10],
            record.get("compliance_score", 0),
            record.get("risk_level", "Medium"),
            record.get("num_findings", 0),
            record.get("analysis_method", "Unknown"),
        ]
    }
    
    # Extract findings from full_analysis
    findings_data = []
    full_analysis = record.get("full_analysis") or {}
    findings = full_analysis.get("findings", [])
    
    for finding in findings:
        findings_data.append({
            "Area": finding.get("area", "Section"),
            "Status": finding.get("status", "Unknown").replace("_", " ").title(),
            "Risk Level": finding.get("risk_level", "Medium"),
            "Issue": finding.get("issue", "N/A"),
            "Suggestion": finding.get("suggestion", "N/A"),
            "Applicable Regulations": "; ".join(finding.get("applicable_regulations", [])),
        })
    
    # Extract citations
    citations_data = []
    citations = full_analysis.get("citations", [])
    
    for cit in citations:
        citations_data.append({
            "Source": cit.get("source", "Unknown"),
            "Title": cit.get("title", "N/A"),
            "URL": cit.get("url", "N/A"),
            "Relevance Score": f"{cit.get('relevance_score', 0):.0%}",
        })
    
    # Create Excel file with multiple sheets
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Summary sheet
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, index=False, sheet_name='Summary')
        
        # Findings sheet
        if findings_data:
            df_findings = pd.DataFrame(findings_data)
            df_findings.to_excel(writer, index=False, sheet_name='Findings')
        
        # Citations sheet
        if citations_data:
            df_citations = pd.DataFrame(citations_data)
            df_citations.to_excel(writer, index=False, sheet_name='Sources')
    
    output.seek(0)
    return output.getvalue()


def _generate_bulk_history_excel(history):
    """
    Generate an Excel file with all analyses as a summary table for bulk export.
    """
    history_data = []
    
    for record in history:
        full_analysis = record.get("full_analysis") or {}
        findings = full_analysis.get("findings", [])
        
        # Count findings by status
        compliant_count = sum(1 for f in findings if f.get("status") == "compliant")
        partial_count = sum(1 for f in findings if f.get("status") == "partially_compliant")
        non_compliant_count = sum(1 for f in findings if f.get("status") == "non_compliant")
        
        timestamp = record.get("timestamp", "")
        try:
            dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
            date_str = dt.strftime("%Y-%m-%d")
        except:
            date_str = timestamp[:10] if timestamp else "Unknown"
        
        history_data.append({
            "Document Name": record.get("document_name", "Unknown"),
            "Domain": record.get("domain", "General"),
            "Analysis Date": date_str,
            "Compliance Score (%)": record.get("compliance_score", 0),
            "Risk Level": record.get("risk_level", "Medium"),
            "Total Findings": record.get("num_findings", 0),
            "Compliant": compliant_count,
            "Partially Compliant": partial_count,
            "Non-Compliant": non_compliant_count,
            "Analysis Method": record.get("analysis_method", "Unknown"),
        })
    
    df = pd.DataFrame(history_data)
    
    # Create Excel with summary
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='All Analyses')
    
    output.seek(0)
    return output.getvalue()


def _generate_analysis_summary_text(record):
    """
    Generate a text summary of the analysis for clipboard export.
    """
    summary = []
    summary.append("=" * 60)
    summary.append("COMPLIANCE ANALYSIS REPORT")
    summary.append("=" * 60)
    summary.append("")
    
    summary.append(f"Document: {record.get('document_name', 'Unknown')}")
    summary.append(f"Domain: {record.get('domain', 'General')}")
    summary.append(f"Analysis Date: {record.get('timestamp', '')[:10]}")
    summary.append("")
    
    summary.append("COMPLIANCE METRICS")
    summary.append("-" * 40)
    summary.append(f"Compliance Score: {record.get('compliance_score', 0)}%")
    summary.append(f"Risk Level: {record.get('risk_level', 'Medium')}")
    summary.append(f"Total Findings: {record.get('num_findings', 0)}")
    summary.append("")
    
    full_analysis = record.get("full_analysis") or {}
    
    # Add executive summary
    exec_summary = full_analysis.get("executive_summary", "")
    if exec_summary:
        summary.append("EXECUTIVE SUMMARY")
        summary.append("-" * 40)
        summary.append(exec_summary)
        summary.append("")
    
    # Add findings
    findings = full_analysis.get("findings", [])
    if findings:
        summary.append("FINDINGS")
        summary.append("-" * 40)
        for i, f in enumerate(findings, 1):
            summary.append(f"\n{i}. {f.get('area', 'Section')}")
            summary.append(f"   Status: {f.get('status', 'Unknown').replace('_', ' ').title()}")
            summary.append(f"   Risk: {f.get('risk_level', 'Medium')}")
            if f.get('issue'):
                summary.append(f"   Issue: {f.get('issue')}")
            if f.get('suggestion'):
                summary.append(f"   Suggestion: {f.get('suggestion')}")
        summary.append("")
    
    # Add missing elements
    missing = full_analysis.get("missing_elements", [])
    if missing:
        summary.append("MISSING REGULATORY ELEMENTS")
        summary.append("-" * 40)
        for m in missing:
            summary.append(f"• {m}")
        summary.append("")
    
    summary.append("=" * 60)
    return "\n".join(summary)


def _get_sample_sop():
    return """STANDARD OPERATING PROCEDURE
Company: Acme Corp
Version: 2.1
Last Updated: January 2025

Section 1: Purpose and Scope
1.1 Purpose
This Standard Operating Procedure establishes the policies and procedures for Acme Corp operations, ensuring compliance with applicable federal, state, and local regulations. This document covers all departments and operational areas.

1.2 Scope
This SOP applies to all employees, contractors, and third-party vendors who perform work on behalf of Acme Corp. It covers data handling, workplace safety, employment practices, financial operations, and environmental compliance.

Section 2: Data Privacy and Protection
2.1 Data Collection Policy
All personal data collected from customers and employees shall be processed in accordance with applicable privacy laws. Data collection must be limited to what is necessary for legitimate business purposes. Consent should be obtained where required.

2.2 Data Storage and Retention
Personal data shall be stored in encrypted databases with access controls. Data retention periods follow industry standards. Records may be kept for business purposes as determined by department heads.

2.3 Data Breach Response
In the event of a suspected data breach, the IT department should be notified. An investigation will be conducted when possible. Affected individuals may be notified if deemed appropriate by management.

Section 3: Workplace Health and Safety
3.1 General Safety Requirements
The company shall provide a safe working environment for all employees. Safety equipment is available upon request. Employees should report hazards to their supervisors.

3.2 Emergency Procedures
Emergency exits are marked throughout the facility. Fire drills are conducted periodically. First aid kits are maintained in common areas.

3.3 Incident Reporting
All workplace incidents should be reported to HR within a reasonable timeframe. Investigation of incidents will be conducted as resources permit.

Section 4: Employment Practices
4.1 Hiring and Onboarding
New employees undergo standard onboarding procedures. Background checks are performed for applicable positions. Equal opportunity principles guide our hiring decisions.

4.2 Compensation and Benefits
Employee compensation follows market standards. Overtime is paid as required by law. Benefits packages are reviewed annually.

4.3 Termination Procedures
Termination of employment follows standard HR procedures. Exit interviews are conducted when possible. Final paychecks are processed according to company policy.

Section 5: Financial Operations
5.1 Accounting Standards
Financial records are maintained using standard accounting practices. Quarterly reviews are performed by the finance team. Tax filings are prepared by the accounting department.

5.2 Expense Management
All business expenses require manager approval. Receipts must be submitted for reimbursement. Corporate credit card usage follows company guidelines.

Section 6: Environmental Compliance
6.1 Waste Management
The company follows applicable waste disposal regulations. Recycling programs are encouraged. Hazardous materials are handled by trained personnel.

6.2 Energy Conservation
Energy efficiency measures are implemented where practical. Equipment is maintained to optimize performance. Conservation initiatives are supported by management.

Section 7: IT Security
7.1 Access Control
System access is granted based on job requirements. Passwords must meet minimum complexity standards. Multi-factor authentication is recommended for sensitive systems.

7.2 Network Security
Firewalls and antivirus software are maintained on all company systems. Regular updates are applied to software and hardware. Security assessments are conducted periodically.

Section 8: Quality Assurance
8.1 Product Quality Standards
Products undergo quality inspection before delivery. Customer feedback is collected and reviewed. Continuous improvement processes are in place.

8.2 Document Control
This SOP is reviewed and updated periodically. Changes require management approval. Previous versions are maintained for reference.
"""


def main():
    # Initialize session state
    if "current_step" not in st.session_state:
        st.session_state["current_step"] = "upload"
    if "last_workflow_step" not in st.session_state:
        st.session_state["last_workflow_step"] = "upload"
    if "nav_page" not in st.session_state:
        st.session_state["nav_page"] = "Workflow"
    
    try:
        render_header()
        render_sidebar()
        
        # Handle navigation based on nav_page radio selection
        nav = st.session_state.get("nav_page", "Workflow")
        current_step = st.session_state.get("current_step", "upload")
        
        if nav == "Feedback":
            # Save current step if leaving workflow
            if current_step not in ["feedback", "history"]:
                st.session_state["last_workflow_step"] = current_step
            st.session_state["current_step"] = "feedback"
        elif nav == "History":
            # Save current step if leaving workflow
            if current_step not in ["feedback", "history"]:
                st.session_state["last_workflow_step"] = current_step
            st.session_state["current_step"] = "history"
        elif nav == "Workflow":
            # Return to workflow - use last saved step or upload
            if current_step in ["feedback", "history"]:
                st.session_state["current_step"] = st.session_state.get("last_workflow_step", "upload")
        
        step = st.session_state.get("current_step", "upload")
        # Only show workflow stepper on workflow pages
        if step not in ["feedback", "history"]:
            render_stepper(step)
        
        # Content
        if step == "upload":
            render_upload_section()
        elif step == "check_document":
            render_check_document()
        elif step == "scrape":
            render_scraping()
        elif step == "analyze":
            render_analysis()
        elif step == "results":
            render_results()
        elif step == "history":
            render_history()
        elif step == "feedback":
            render_feedback()
        else:
            st.error(f"Unknown step: {step}")
    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
        st.exception(e)


if __name__ == "__main__":
    main()
